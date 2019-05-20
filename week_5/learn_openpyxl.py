# -*- coding: utf-8 -*-
# @Time : 2019/3/25 22:32

# openpyxl模块的学习&编写类

# 创建Excel文件的模块 workbook
# 读写Excel文件的模块 Load_workbook

# 操作Excel：workbook sheet cell
from openpyxl import workbook
# wb = workbook.Workbook()  # 新建一个Excel
# wb.create_sheet('sheet_hanfan') # 创建一个你自己命名的表单sheet
# wb.save('python_hanf.xlsx') # 另存为，默认保存为当前路径，Excel后缀名必须为xlsx

# 开始读写excel
# 第一步 打开工作簿
# from openpyxl import load_workbook
# wb=load_workbook('python_hanf.xlsx')  # 打开Excel
# #第二步 定位工作表
# sheet = wb['sheet_hanfan']  #定位到工作表sheet
# #第三步 定位单元格  根据行列坐标定位单元格取值
# res= sheet.cell(3,3).value
# print('res的值是:{}，类型是{}'.format(res,type(res)))  #------> res的值是:[1,2,3]，类型是<class 'str'>
# # 结论：数字类型读出后还是int，其他数据类型读出都是字符串str
#
# res = eval(sheet.cell(3,3).value)
# print('res的值是:{}，类型是{}'.format(res,type(res)))  # ----->res的值是:[1, 2, 3]，类型是<class 'list'>
# #结论：eval()可以把数据转换成Python原本可以识别的数据类型，比如：原本是列表，打印出来还是列
#
#
# # 写入值、保存  两种方式(新增与修改都是如下方法)
# # 写值时，记得关闭工作簿，不然会报错
# sheet.cell(5,1).value='白日依山尽'  # 方法一
# sheet.cell(5,2,'黄河入海流')   # 方法二
# # 保存工作簿
# wb.save('python_hanf.xlsx')
# # 操作完毕后记得关闭
# wb.close()

# ==================================================================================

from openpyxl import load_workbook
wb = load_workbook('python_hanf.xlsx')
sheet = wb['sheet_hanfan']

print(sheet.max_row)  #获取最大行
print(sheet.max_column)  #获取最大列
# 打印出工作表的所有内容
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        if sheet.cell(i,j).value != None:  # 单元格为空的时候，读取到的是None;此句判断是否为空，为空则不执行下面语句
            print(sheet.cell(i,j).value)


