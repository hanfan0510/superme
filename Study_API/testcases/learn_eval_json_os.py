# _*_ coding: utf-8  _*_ 
# @Time : 2019-04-19 01:18
# 2019.4.13号的视频，编号46


# eval和json区别

'''练习打开工作表'''
'''练习eval的使用方法'''

from openpyxl import load_workbook

'''工作表单元格(2，4)处的内容为：{'mobilephone': '15011096051', 'pwd': '123456'}'''

wb = load_workbook('cases.xlsx')  # 打开工作表
sheet = wb['login']  # 定位工作表sheet
res = sheet.cell(2,4).value  # 定位单元格

print(res)  # {'mobilephone': '15011096051', 'pwd': '123456'}   从Excel取出来的是一个字符串
print(type(res))  # <class 'str'>

'''eval处理，还原res数据类型'''
res2 = eval(res)  # python类型的字符串转换为字典
print(res2)  # {'mobilephone': '15011096051', 'pwd': '123456'}
print(res2['pwd'])
print(type(res2))  # <class 'dict'>

print()  # 仅仅为了换行

'''eval和json的区别：
eval是python语法
json语法：
{"status":1,"code":"10001","data":null,"msg":"登录成功"}
1、值必须用双引号，python用单引号
2、空用null表示，python用None
3、布尔类型的写为true和false，python用True和False

根据请求入参判断是用eval还是json.loads转换
'''

# 举例：
import json

param = '{"status":true,"code":"10001","data":null,"msg":"登录成功"}'  # json格式的字符串  注意看true和null变化
p = json.loads(param)  # 将json转换为python的数据类型的字典
print(type(p))  # <class 'dict'>
print(p)  # {'status': True, 'code': '10001', 'data': None, 'msg': '登录成功'}   注意看true和null变化
print(p['msg'])




'''
学习os： 查找文件绝对路径
'''

import os

case_dir = os.path.abspath(__file__)
print(case_dir)  # /Users/hanfan/PycharmProjects/class_base/Study_API/learn_eval_json_os.py

base_dir = os.path.dirname(os.path.dirname(case_dir))
print(base_dir)  # /Users/hanfan/PycharmProjects/class_base

case_file = os.path.join(base_dir,'task','task_0410.py')   # 路径合并
print(case_file)   # /Users/hanfan/PycharmProjects/class_base/task/task_0410.py
